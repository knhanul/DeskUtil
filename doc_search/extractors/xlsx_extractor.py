"""XLSX 추출기"""
from pathlib import Path
import openpyxl

class XLSXExtractor:
    """XLSX 텍스트 추출기"""
    
    def extract_text(self, file_path: str) -> str:
        """XLSX에서 텍스트 추출"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text += str(cell) + " "
            
            workbook.close()
            return text.strip()
        except:
            return ""
